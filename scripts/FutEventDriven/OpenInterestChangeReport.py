import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
import datetime
import warnings

import data.ConstantData.future_basic_information as ConstFutBasic
import data.SQL.extract_data_from_postgre as ExtractDataPostgre
import data.Process.fut_data_process as DataProcessFut
from scripts import LoggingConfig
import time

logger = LoggingConfig.Config().get_config()
warnings.filterwarnings("ignore")

group_cffex_stock = ['IC.CFE', 'IF.CFE', 'IH.CFE', 'IM.CFE']
group_monthly_roll = ['AL.SHF', 'CU.SHF', 'SP.SHF', 'PB.SHF', 'PG.DCE', 'EB.DCE', 'NI.SHF', 'SN.SHF', 'ZN.SHF',
                      'SS.SHF']
group_quarterly_roll = ['TF.CFE', 'T.CFE', 'TS.CFE']
group_three_times_yearly_roll = ['PK.CZC', 'TA.CZC', 'AP.CZC', 'RB.SHF', 'HC.SHF', 'C.DCE', 'M.DCE', 'A.DCE', 'BU.SHF',
                                 'SR.CZC', 'JM.DCE', 'JD.DCE', 'CS.DCE', 'FG.CZC', 'FU.SHF', 'SM.CZC', 'I.DCE',
                                 'SF.CZC', 'CJ.CZC', 'J.DCE', 'EG.DCE', 'LH.DCE', 'P.DCE', 'L.DCE', 'V.DCE', 'PP.DCE',
                                 'RU.SHF', 'Y.DCE', 'ZC.CZC', 'UR.CZC', 'CF.CZC', 'MA.CZC', 'OI.CZC', 'RM.CZC',
                                 'SA.CZC', 'PF.CZC']
window_monthly_roll = 12
window_yoy_monthly_roll = 3
window_non_monthly_roll = 9
window_yoy_non_monthly_roll = 4


############################
# calculate the database

def calculate_one_fut_code_oi_change_rate_indicator(tickers, start_date, end_date, tickers_list_delist_date_dict=None):
    # 所有期货合约上市与合约到期时间
    if tickers_list_delist_date_dict is None:
        tickers_list_delist_date_df = ExtractDataPostgre.get_list_date_and_delist_date_of_instrument()
        tickers_list_delist_date_dict = {
            tickers_list_delist_date_df.loc[i, 'ts_code']: tickers_list_delist_date_df.loc[i, 'delist_date'] for i in
            range(len(tickers_list_delist_date_df))}
    logger.info("tickers_list_delist_date_dict succeed")

    # 交易所的交易日历
    trading_calendar = np.array(ExtractDataPostgre.get_trading_calendar(exchange=tickers.split('.')[-1]).iloc[:, 0])
    logger.info("trading_calendar succeed")

    # future上市未满一年取上市一年后作为start date
    start_date_revise = DataProcessFut.adjust_start_date(start_date, tickers, threshold=1, delta_time=365)
    logger.info("start_date_revise succeed")

    # 提取区间范围内 gsci主力与次主力时序 及主力合约open interest matrix
    # 提取fut code所有合约及其open interest时间序列
    tickers_all = [key for key, value in ExtractDataPostgre.get_code_instrument_mapping().items() if value == tickers]
    logger.info("Load all instrument of the future code")
    df_oi = ExtractDataPostgre.get_future_daily_data_sm(tickers=tickers_all, start_date=start_date_revise,
                                                        end_date=end_date, key_word='oi', as_matrix=True)
    logger.info("Load open interest ts data succeed")
    df_close = ExtractDataPostgre.get_future_daily_data_sm(tickers=tickers_all, start_date=start_date_revise,
                                                           end_date=end_date, key_word='close', as_matrix=True)
    logger.info("Load close price ts data succeed")

    # 求和，得到fut code总持仓的时间序列
    open_interest_daily = df_oi.apply(np.nansum, axis=1).reset_index()
    open_interest_daily.columns = ['trade_date', 'oi_all']
    logger.info("Calculated daily total open interest succeed")

    # 计算每一天主力、次主力、次次主力的占比； 主力合约open interest 占比matrix
    df_oi = df_oi.reset_index()
    df_close = df_close.reset_index()
    df_oi_all = pd.merge(left=df_oi, right=open_interest_daily, how='left', on='trade_date')
    df_oi_pct = df_oi.copy()
    for i in range(1, len(df_oi_pct.columns)):
        df_oi_pct.loc[:, df_oi_pct.columns[i]] = df_oi_all.iloc[:, i] / df_oi_all['oi_all']
    logger.info("Calculated daily open interest pct succeed")

    # melt oi indicator

    df_oi_pct = df_oi_pct.melt('trade_date')
    df_oi_pct.columns = ['tradingday', 'code', 'oi_rate']
    df_oi = df_oi.melt('trade_date')
    df_oi.columns = ['tradingday', 'code', 'oi']
    df_close = df_close.melt('trade_date')
    df_close.columns = ['tradingday', 'code', 'close']
    logger.info("data oi, oi_pct, close melt")

    # 根据到期日roll的时间序列
    logger.info("start generate delist roll index dataframe")
    data = ExtractDataPostgre.get_delist_date_roll_df(tickers=tickers, start_date=start_date_revise, end_date=end_date)
    for i in range(len(data)):
        if i == 0:
            try:
                data.loc[i, 'second_main_instrumentid'] = DataProcessFut.get_next_main_contract(
                    data.loc[i, 'current_main_instrumentid'])
                data.loc[i, 'third_main_instrumentid'] = DataProcessFut.get_next_main_contract(
                    data.loc[i, 'second_main_instrumentid'])
            except Exception as e:
                data.loc[i, 'second_main_instrumentid'] = np.nan
                data.loc[i, 'third_main_instrumentid'] = np.nan
        else:
            if data.loc[i, 'current_main_instrumentid'] == data.loc[i - 1, 'current_main_instrumentid']:
                data.loc[i, 'second_main_instrumentid'] = data.loc[i - 1, 'second_main_instrumentid']
                data.loc[i, 'third_main_instrumentid'] = data.loc[i - 1, 'third_main_instrumentid']
            else:
                try:
                    data.loc[i, 'second_main_instrumentid'] = DataProcessFut.get_next_main_contract(
                        data.loc[i, 'current_main_instrumentid'])
                    data.loc[i, 'third_main_instrumentid'] = DataProcessFut.get_next_main_contract(
                        data.loc[i, 'second_main_instrumentid'])
                except Exception as e:
                    data.loc[i, 'second_main_instrumentid'] = np.nan
                    data.loc[i, 'third_main_instrumentid'] = np.nan
    logger.info("end generate delist roll index dataframe")

    # merge oi rate
    data_all = pd.merge(left=data, right=df_oi_pct, how='left', left_on=['tradingday', 'current_main_instrumentid'],
                        right_on=['tradingday', 'code'])
    data_all = data_all[['main_contract_code', 'tradingday', 'current_main_instrumentid', 'second_main_instrumentid',
                         'third_main_instrumentid', 'oi_rate']]
    data_all.columns = ['main_contract_code', 'tradingday', 'current_main_instrumentid', 'second_main_instrumentid',
                        'third_main_instrumentid', 'main_oi_rate']
    data_all = pd.merge(left=data_all, right=df_oi_pct, how='left', left_on=['tradingday', 'second_main_instrumentid'],
                        right_on=['tradingday', 'code'])
    data_all = data_all[['main_contract_code', 'tradingday', 'current_main_instrumentid', 'second_main_instrumentid',
                         'third_main_instrumentid', 'main_oi_rate', 'oi_rate']]
    data_all.columns = ['main_contract_code', 'tradingday', 'current_main_instrumentid', 'second_main_instrumentid',
                        'third_main_instrumentid', 'main_oi_rate', 'second_oi_rate']
    data_all = pd.merge(left=data_all, right=df_oi_pct, how='left', left_on=['tradingday', 'third_main_instrumentid'],
                        right_on=['tradingday', 'code'])
    data_all = data_all[['main_contract_code', 'tradingday', 'current_main_instrumentid', 'second_main_instrumentid',
                         'third_main_instrumentid', 'main_oi_rate', 'second_oi_rate', 'oi_rate']]
    data_all.columns = ['main_contract_code', 'tradingday', 'current_main_instrumentid', 'second_main_instrumentid',
                        'third_main_instrumentid', 'main_oi_rate', 'second_oi_rate', 'third_oi_rate']
    logger.info('merge oi rate')
    # merge oi
    data_all = pd.merge(left=data_all, right=df_oi, how='left', left_on=['tradingday', 'current_main_instrumentid'],
                        right_on=['tradingday', 'code'])
    data_all = data_all[['main_contract_code', 'tradingday', 'current_main_instrumentid',
                         'second_main_instrumentid', 'third_main_instrumentid', 'main_oi_rate',
                         'second_oi_rate', 'third_oi_rate', 'oi']]
    data_all.columns = ['main_contract_code', 'tradingday', 'current_main_instrumentid',
                        'second_main_instrumentid', 'third_main_instrumentid', 'main_oi_rate',
                        'second_oi_rate', 'third_oi_rate', 'main_oi']
    data_all = pd.merge(left=data_all, right=df_oi, how='left', left_on=['tradingday', 'second_main_instrumentid'],
                        right_on=['tradingday', 'code'])
    data_all = data_all[['main_contract_code', 'tradingday', 'current_main_instrumentid',
                         'second_main_instrumentid', 'third_main_instrumentid', 'main_oi_rate',
                         'second_oi_rate', 'third_oi_rate', 'main_oi', 'oi']]
    data_all.columns = ['main_contract_code', 'tradingday', 'current_main_instrumentid',
                        'second_main_instrumentid', 'third_main_instrumentid', 'main_oi_rate',
                        'second_oi_rate', 'third_oi_rate', 'main_oi', 'second_oi']
    data_all = pd.merge(left=data_all, right=df_oi, how='left', left_on=['tradingday', 'third_main_instrumentid'],
                        right_on=['tradingday', 'code'])
    data_all = data_all[['main_contract_code', 'tradingday', 'current_main_instrumentid',
                         'second_main_instrumentid', 'third_main_instrumentid', 'main_oi_rate',
                         'second_oi_rate', 'third_oi_rate', 'main_oi', 'second_oi', 'oi']]
    data_all.columns = ['main_contract_code', 'tradingday', 'current_main_instrumentid',
                        'second_main_instrumentid', 'third_main_instrumentid', 'main_oi_rate',
                        'second_oi_rate', 'third_oi_rate', 'main_oi', 'second_oi', 'third_oi']
    logger.info('merge oi')
    # merge close
    data_all = pd.merge(left=data_all, right=df_close, how='left', left_on=['tradingday', 'current_main_instrumentid'],
                        right_on=['tradingday', 'code'])
    data_all = data_all[['main_contract_code', 'tradingday', 'current_main_instrumentid',
                         'second_main_instrumentid', 'third_main_instrumentid', 'main_oi_rate',
                         'second_oi_rate', 'third_oi_rate', 'main_oi', 'second_oi', 'third_oi',
                         'close']]
    data_all.columns = ['main_contract_code', 'tradingday', 'current_main_instrumentid',
                        'second_main_instrumentid', 'third_main_instrumentid', 'main_oi_rate',
                        'second_oi_rate', 'third_oi_rate', 'main_oi', 'second_oi', 'third_oi',
                        'main_close']
    data_all = pd.merge(left=data_all, right=df_close, how='left', left_on=['tradingday', 'second_main_instrumentid'],
                        right_on=['tradingday', 'code'])
    data_all = data_all[['main_contract_code', 'tradingday', 'current_main_instrumentid',
                         'second_main_instrumentid', 'third_main_instrumentid', 'main_oi_rate',
                         'second_oi_rate', 'third_oi_rate', 'main_oi', 'second_oi', 'third_oi',
                         'main_close', 'close']]
    data_all.columns = ['main_contract_code', 'tradingday', 'current_main_instrumentid',
                        'second_main_instrumentid', 'third_main_instrumentid', 'main_oi_rate',
                        'second_oi_rate', 'third_oi_rate', 'main_oi', 'second_oi', 'third_oi',
                        'main_close', 'second_close']
    data_all = pd.merge(left=data_all, right=df_close, how='left', left_on=['tradingday', 'third_main_instrumentid'],
                        right_on=['tradingday', 'code'])
    data_all = data_all[['main_contract_code', 'tradingday', 'current_main_instrumentid',
                         'second_main_instrumentid', 'third_main_instrumentid', 'main_oi_rate',
                         'second_oi_rate', 'third_oi_rate', 'main_oi', 'second_oi', 'third_oi',
                         'main_close', 'second_close', 'close']]
    data_all.columns = ['main_contract_code', 'tradingday', 'current_main_instrumentid',
                        'second_main_instrumentid', 'third_main_instrumentid', 'main_oi_rate',
                        'second_oi_rate', 'third_oi_rate', 'main_oi', 'second_oi', 'third_oi',
                        'main_close', 'second_close', 'third_close']
    logger.info('merge close price')

    # 计算合约到期日和到期时长
    main_delist_index = 0
    second_delist_index = 0
    third_delist_index = 0

    for i in range(len(data_all)):
        trading_day_index = np.where(trading_calendar == data_all.loc[i, 'tradingday'])[0][0]
        if i == 0:

            # delist date
            try:
                data_all.loc[i, 'main_delist'] = tickers_list_delist_date_dict[
                    data_all.loc[i, 'current_main_instrumentid']]
            except Exception as e:
                data_all.loc[i, 'main_delist'] = np.nan
            try:
                data_all.loc[i, 'second_delist'] = tickers_list_delist_date_dict[
                    data_all.loc[i, 'second_main_instrumentid']]
            except Exception as e:
                data_all.loc[i, 'second_delist'] = np.nan
            try:
                data_all.loc[i, 'third_delist'] = tickers_list_delist_date_dict[
                    data_all.loc[i, 'third_main_instrumentid']]
            except Exception as e:
                data_all.loc[i, 'third_delist'] = np.nan

            # find calendar index
            try:
                main_delist_index = np.where(trading_calendar == data_all.loc[i, 'main_delist'])[0][0]
            except Exception as e:
                main_delist_index = None
            try:
                second_delist_index = np.where(trading_calendar == data_all.loc[i, 'second_delist'])[0][0]
            except Exception as e:
                second_delist_index = None
            try:
                third_delist_index = np.where(trading_calendar == data_all.loc[i, 'third_delist'])[0][0]
            except Exception as e:
                third_delist_index = 0

            # calculate TTM
            if pd.isna(data_all.loc[i, 'main_delist']):
                data_all.loc[i, 'main_TTM'] = np.nan
            else:
                if main_delist_index is None:
                    data_all.loc[i, 'main_TTM'] = round(
                        (data_all.loc[i, 'main_delist'] - data_all.loc[i, 'tradingday']).days * 5 / 7)
                else:
                    data_all.loc[i, 'main_TTM'] = main_delist_index - trading_day_index
            if pd.isna(data_all.loc[i, 'second_delist']):
                data_all.loc[i, 'second_TTM'] = np.nan
            else:
                if second_delist_index is None:
                    data_all.loc[i, 'second_TTM'] = round(
                        (data_all.loc[i, 'second_delist'] - data_all.loc[i, 'tradingday']).days * 5 / 7)
                else:
                    data_all.loc[i, 'second_TTM'] = second_delist_index - trading_day_index
            if pd.isna(data_all.loc[i, 'third_delist']):
                data_all.loc[i, 'third_TTM'] = np.nan
            else:
                if third_delist_index is None:
                    data_all.loc[i, 'third_TTM'] = round(
                        (data_all.loc[i, 'third_delist'] - data_all.loc[i, 'tradingday']).days * 5 / 7)
                else:
                    data_all.loc[i, 'third_TTM'] = third_delist_index - trading_day_index
        else:
            if data_all.loc[i, 'current_main_instrumentid'] == data_all.loc[i - 1, 'current_main_instrumentid']:
                # copy delist before
                data_all.loc[i, 'main_delist'] = data_all.loc[i - 1, 'main_delist']
                data_all.loc[i, 'second_delist'] = data_all.loc[i - 1, 'second_delist']
                data_all.loc[i, 'third_delist'] = data_all.loc[i - 1, 'third_delist']

                # calculate TTM
                if pd.isna(data_all.loc[i, 'main_delist']):
                    data_all.loc[i, 'main_TTM'] = np.nan
                else:
                    if main_delist_index is None:
                        data_all.loc[i, 'main_TTM'] = round(
                            (data_all.loc[i, 'main_delist'] - data_all.loc[i, 'tradingday']).days * 5 / 7)
                    else:
                        data_all.loc[i, 'main_TTM'] = main_delist_index - trading_day_index
                if pd.isna(data_all.loc[i, 'second_delist']):
                    data_all.loc[i, 'second_TTM'] = np.nan
                else:
                    if second_delist_index is None:
                        data_all.loc[i, 'second_TTM'] = round(
                            (data_all.loc[i, 'second_delist'] - data_all.loc[i, 'tradingday']).days * 5 / 7)
                    else:
                        data_all.loc[i, 'second_TTM'] = second_delist_index - trading_day_index
                if pd.isna(data_all.loc[i, 'third_delist']):
                    data_all.loc[i, 'third_TTM'] = np.nan
                else:
                    if third_delist_index is None:
                        data_all.loc[i, 'third_TTM'] = round(
                            (data_all.loc[i, 'third_delist'] - data_all.loc[i, 'tradingday']).days * 5 / 7)
                    else:
                        data_all.loc[i, 'third_TTM'] = third_delist_index - trading_day_index
            else:
                # delist date
                try:
                    data_all.loc[i, 'main_delist'] = tickers_list_delist_date_dict[
                        data_all.loc[i, 'current_main_instrumentid']]
                except Exception as e:
                    data_all.loc[i, 'main_delist'] = np.nan
                try:
                    data_all.loc[i, 'second_delist'] = tickers_list_delist_date_dict[
                        data_all.loc[i, 'second_main_instrumentid']]
                except Exception as e:
                    data_all.loc[i, 'second_delist'] = np.nan
                try:
                    data_all.loc[i, 'third_delist'] = tickers_list_delist_date_dict[
                        data_all.loc[i, 'third_main_instrumentid']]
                except Exception as e:
                    data_all.loc[i, 'third_delist'] = np.nan

                # calendar index
                try:
                    main_delist_index = np.where(trading_calendar == data_all.loc[i, 'main_delist'])[0][0]
                except Exception as e:
                    main_delist_index = None
                try:
                    second_delist_index = np.where(trading_calendar == data_all.loc[i, 'second_delist'])[0][0]
                except Exception as e:
                    second_delist_index = None
                try:
                    third_delist_index = np.where(trading_calendar == data_all.loc[i, 'third_delist'])[0][0]
                except Exception as e:
                    third_delist_index = None

                # calculate TTM
                if pd.isna(data_all.loc[i, 'main_delist']):
                    data_all.loc[i, 'main_TTM'] = np.nan
                else:
                    if main_delist_index is None:
                        data_all.loc[i, 'main_TTM'] = round(
                            (data_all.loc[i, 'main_delist'] - data_all.loc[i, 'tradingday']).days * 5 / 7)
                    else:
                        data_all.loc[i, 'main_TTM'] = main_delist_index - trading_day_index
                if pd.isna(data_all.loc[i, 'second_delist']):
                    data_all.loc[i, 'second_TTM'] = np.nan
                else:
                    if second_delist_index is None:
                        data_all.loc[i, 'second_TTM'] = round(
                            (data_all.loc[i, 'second_delist'] - data_all.loc[i, 'tradingday']).days * 5 / 7)
                    else:
                        data_all.loc[i, 'second_TTM'] = second_delist_index - trading_day_index
                if pd.isna(data_all.loc[i, 'third_delist']):
                    data_all.loc[i, 'third_TTM'] = np.nan
                else:
                    if third_delist_index is None:
                        data_all.loc[i, 'third_TTM'] = round(
                            (data_all.loc[i, 'third_delist'] - data_all.loc[i, 'tradingday']).days * 5 / 7)
                    else:
                        data_all.loc[i, 'third_TTM'] = third_delist_index - trading_day_index

    logger.info('calculate delist date and TTM')

    # 计算合约需要rolling的速度，spread，spread pct
    data_all['main_OI_speed'] = data_all['main_oi'] / (data_all['main_TTM'] + 1)
    data_all['second_OI_speed'] = data_all['second_oi'] / (data_all['second_TTM'] + 1)
    data_all['third_OI_speed'] = data_all['third_oi'] / (data_all['third_TTM'] + 1)
    data_all['spread_01'] = data_all['main_close'] - data_all['second_close']
    data_all['spread_12'] = data_all['second_close'] - data_all['third_close']
    data_all['spread_pct_01'] = (data_all['main_close'] - data_all['second_close']) / data_all['main_close'] * 252 / (
            data_all['main_TTM'] + 1)
    data_all['spread_pct_12'] = (data_all['second_close'] - data_all['third_close']) / data_all[
        'second_close'] * 252 / (data_all['second_TTM'] + 1)
    logger.info('calculate oi speed, spread, spread_pct')

    return data_all


def oi_change_rate_indicator(tickers_list, start_date, end_date, file_path, daily_index=False):
    result_all = pd.DataFrame()
    tickers_list_delist_date_df = ExtractDataPostgre.get_list_date_and_delist_date_of_instrument()
    tickers_list_delist_date_dict = {
        tickers_list_delist_date_df.loc[i, 'ts_code']: tickers_list_delist_date_df.loc[i, 'delist_date'] for i in
        range(len(tickers_list_delist_date_df))}
    for i in range(len(tickers_list)):
        tickers = tickers_list[i]
        try:
            logger.info(f"start tickers: {tickers}")
            data_all = calculate_one_fut_code_oi_change_rate_indicator(tickers, start_date, end_date,
                                                                       tickers_list_delist_date_dict=tickers_list_delist_date_dict)
            file_path_save = file_path + end_date + '\\'
            if not os.path.exists(file_path_save):
                os.makedirs(file_path_save)
            data_all.to_csv(file_path + end_date + '\\' + tickers + ".csv")
            # data_all=pd.read_csv(file_path + end_date + '\\' + tickers + ".csv",index_col=0)
            result_all = pd.concat([result_all, data_all])

            if daily_index:
                result_all = result_all[
                    result_all['tradingday'] == datetime.date(int(end_date[:4]), int(end_date[5:7]),
                                                              int(end_date[-2:]))]
            result_all.to_csv(file_path + 'all\\' + end_date + '.csv')

            logger.info(f"end tickers: {tickers},{len(result_all)}")
        except Exception as e:
            logger.info(e)
        time.sleep(5)
    return result_all


def oi_change_rate_all(
        file_path='C:\\Users\\jason.huang\\research\\scripts_working\\OI_change_rate\\daily\\2022-09-05\\'):
    data_all = pd.DataFrame()
    for file_name in os.listdir(file_path):
        file_location = file_path + file_name
        data = pd.read_csv(file_location, index_col=0)
        data = data.iloc[-2:, :]
        data_all = pd.concat([data_all, data])
    data_all.to_csv('111.csv')


##############################
# calculate oi change daily table

def determine_main_contract(df):
    for i in range(len(df)):
        date_str = df.iloc[i, 0][2:4] + df.iloc[i, 0][5:7]
        code1 = df.iloc[i, 1].split('.')[0][-4:]
        if int(date_str) < int(code1):
            return 'current', 'second'
        else:
            return 'second', 'third'


def oi_change_rate_table_one_ticker(tickers, date, data):
    result = pd.DataFrame()
    result.loc[0, 'Code'] = tickers
    result.loc[0, 'Name'] = ConstFutBasic.fut_code_to_chinese_name_mapping[tickers][:-2]
    # if \
    # ConstFutBasic.fut_code_to_chinese_name_mapping[tickers][-2:] == '主力' else \
    # ConstFutBasic.fut_code_to_chinese_name_mapping[tickers]
    df_all = data[data['main_contract_code'] == tickers]
    df_all['current_spread_pct'] = (-df_all['current_close'] + df_all['second_close']) / df_all['current_close']
    df_all['second_spread_pct'] = (-df_all['second_close'] + df_all['third_close']) / df_all['second_close']
    df_all['current_spread_pct_lag_1'] = df_all['current_spread_pct'].shift(-1)
    df_all['current_spread_pct_lag_2'] = df_all['current_spread_pct'].shift(-2)
    df_all['current_spread_pct_lag_3'] = df_all['current_spread_pct'].shift(-3)
    df_all['current_spread_pct_lag_4'] = df_all['current_spread_pct'].shift(-4)

    df_all['second_spread_pct_lag_1'] = df_all['second_spread_pct'].shift(-1)
    df_all['second_spread_pct_lag_2'] = df_all['second_spread_pct'].shift(-2)
    df_all['second_spread_pct_lag_3'] = df_all['second_spread_pct'].shift(-3)
    df_all['second_spread_pct_lag_4'] = df_all['second_spread_pct'].shift(-4)

    df_all = df_all.sort_values('tradingday').reset_index(drop=True)
    df_obj = df_all[df_all['tradingday'] == date].reset_index(drop=True)

    if tickers in group_cffex_stock:
        str1, str2 = 'current', 'second'
    else:
        str1, str2 = determine_main_contract(df_obj.iloc[:, 1:5])

    if tickers in group_monthly_roll:
        window = window_monthly_roll
        window_yoy = window_yoy_monthly_roll
    else:
        window = window_non_monthly_roll
        window_yoy = window_yoy_non_monthly_roll

    result.loc[0, 'Roll_Out'] = df_obj.loc[0, str1 + '_main_instrumentid']
    result.loc[0, 'Roll_In'] = df_obj.loc[0, str2 + '_main_instrumentid']
    result.loc[0, 'Roll_TTM'] = df_obj.loc[0, str1 + '_TTM']
    month = result.loc[0, 'Roll_Out'].split('.')[0][-2:]
    ttm = df_obj.loc[0, str1 + '_TTM']

    df_ttm = df_all[df_all['current_TTM'] == ttm]
    df_ttm = df_ttm[df_ttm['tradingday'] < date].reset_index(drop=True)
    df_ttm = df_ttm[
        ['main_contract_code', 'tradingday'] + [x for x in df_ttm.columns if 'current' in x or 'second' in x]]
    df_ttm.columns = [x.replace('current', 'first') if 'current' in x else x for x in df_ttm.columns]
    df_ttm = df_ttm[['main_contract_code', 'tradingday', 'first_main_instrumentid',
                     'second_main_instrumentid', 'first_oi_rate', 'second_oi_rate',
                     'first_oi', 'second_oi', 'first_close', 'second_close', 'first_delist',
                     'second_delist', 'first_TTM', 'second_TTM', 'first_OI_speed',
                     'second_OI_speed', 'first_spread_pct',
                     'first_spread_pct_lag_1', 'first_spread_pct_lag_2',
                     'first_spread_pct_lag_3', 'first_spread_pct_lag_4']]
    df_ttm.loc[:, 'month'] = df_ttm.loc[:, 'first_main_instrumentid'].apply(lambda x: x.split('.')[0][-2:])

    df_ttm2 = df_all[df_all['second_TTM'] == ttm]
    df_ttm2 = df_ttm2[df_ttm2['tradingday'] < date].reset_index(drop=True)
    df_ttm2 = df_ttm2[
        ['main_contract_code', 'tradingday'] + [x for x in df_ttm2.columns if 'second' in x or 'third' in x]]
    df_ttm2.columns = [x.replace('second', 'first') if 'second' in x else x for x in df_ttm2.columns]
    df_ttm2.columns = [x.replace('third', 'second') if 'third' in x else x for x in df_ttm2.columns]
    df_ttm2.loc[:, 'month'] = df_ttm2.loc[:, 'first_main_instrumentid'].apply(lambda x: x.split('.')[0][-2:])
    df_ttm = pd.concat([df_ttm, df_ttm2])
    df_ttm = df_ttm.sort_values('tradingday').reset_index(drop=True)

    # df_ttm = df_all[df_all[str1 + '_TTM'] == ttm]
    # df_ttm.loc[:, 'month'] = df_ttm.loc[:, str1 + '_main_instrumentid'].apply(lambda x: x.split('.')[0][-2:])
    # df_ttm = df_ttm[df_ttm['tradingday'] < date].reset_index(drop=True)

    df1 = df_ttm.iloc[-1 * window:, :]
    df2 = df_ttm[df_ttm['month'] == month].reset_index(drop=True)
    df2 = df2.iloc[-1 * window_yoy:, :]

    str1_revise = 'first'

    result.loc[0, 'OI_pct'] = df_obj.loc[0, str1 + '_oi_rate'] - np.nanmean(df1.loc[:, str1_revise + '_oi_rate'])
    result.loc[0, 'OI_yoy_pct'] = df_obj.loc[0, str1 + '_oi_rate'] - np.nanmean(df2.loc[:, str1_revise + '_oi_rate'])
    result.loc[0, 'OI_speed_pct'] = df_obj.loc[0, str1 + '_OI_speed'] - np.nanmean(
        df1.loc[:, str1_revise + '_OI_speed'])
    result.loc[0, 'OI_speed_yoy_pct'] = df_obj.loc[0, str1 + '_OI_speed'] - np.nanmean(
        df2.loc[:, str1_revise + '_OI_speed'])
    result.loc[0, 'signal'] = np.nansum(
        [np.sign(result.loc[0, 'OI_pct']), np.sign(result.loc[0, 'OI_yoy_pct']), np.sign(
            result.loc[0, 'OI_speed_pct']), np.sign(result.loc[0, 'OI_speed_yoy_pct'])])
    result.loc[0, 'roll_yield'] = np.nanmedian(
        [np.nanmedian(df1.loc[:, str1_revise + '_spread_pct_lag_1']),
         np.nanmedian(df1.loc[:, str1_revise + '_spread_pct_lag_2']),
         np.nanmedian(df1.loc[:, str1_revise + '_spread_pct_lag_3']),
         np.nanmedian(df1.loc[:, str1_revise + '_spread_pct_lag_4'])])
    result.loc[0, 'roll_yield_yoy'] = np.nanmean(
        [np.nanmedian(df2.loc[:, str1_revise + '_spread_pct_lag_1']),
         np.nanmedian(df2.loc[:, str1_revise + '_spread_pct_lag_2']),
         np.nanmedian(df2.loc[:, str1_revise + '_spread_pct_lag_3']),
         np.nanmedian(df2.loc[:, str1_revise + '_spread_pct_lag_4'])])

    return result


def oi_change_rate_table(date,
                         file_path='C:\\Users\\jason.huang\\research\\scripts_working\\OI_change_rate\\daily\\all\\'):
    data = pd.read_csv("C:\\Users\\jason.huang\\research\\scripts_working\\OI_change_rate\\all.csv", index_col=0)
    for i in range(len(os.listdir(file_path))):
        file_location = file_path + os.listdir(file_path)[i]
        df = pd.read_csv(file_location, index_col=0)
        data = pd.concat([data, df])
    data.columns = ['main_contract_code', 'tradingday', 'current_main_instrumentid',
                    'second_main_instrumentid', 'third_main_instrumentid', 'current_oi_rate',
                    'second_oi_rate', 'third_oi_rate', 'current_oi', 'second_oi', 'third_oi',
                    'current_close', 'second_close', 'third_close', 'current_delist',
                    'second_delist', 'third_delist', 'current_TTM', 'second_TTM', 'third_TTM',
                    'current_OI_speed', 'second_OI_speed', 'third_OI_speed', 'spread_01',
                    'spread_12', 'spread_pct_01', 'spread_pct_12']
    result_all = pd.DataFrame()
    for tickers in ConstFutBasic.fut_code_list:
        try:
            result_one_ticker = oi_change_rate_table_one_ticker(tickers, date, data)
            result_all = pd.concat([result_all, result_one_ticker])
        except Exception as e:
            print(tickers + ' is not ok')

    result_all = result_all.sort_values('Roll_TTM')

    workbook = load_workbook(
        filename="C:\\Users\\jason.huang\\research\\scripts_working\\OI_change_rate\\table\\format.xlsx")
    sheet = workbook.active
    for j in range(len(result_all.columns)):
        sheet.cell(row=1, column=j + 1).value = result_all.columns[j]

        for i in range(len(result_all)):
            sheet.cell(row=i + 2, column=j + 1).value = result_all.iloc[i, j]
    workbook.save(
        filename="C:\\Users\\jason.huang\\research\\scripts_working\\OI_change_rate\\table\\" + date + ".xlsx")
    result_all.to_csv(
        "C:\\Users\\jason.huang\\research\\scripts_working\\OI_change_rate\\table\\" + date + "_OI_change.csv",
        encoding='utf-8-sig', index=0)


###############################
# OI chg factor

def oi_change_rate_factor():
    data = pd.read_csv("C:\\Users\\jason.huang\\research\\factor_data\\OI_change\\all.csv", index_col=0)
    data.columns = ['main_contract_code', 'tradingday', 'current_main_instrumentid',
                    'second_main_instrumentid', 'third_main_instrumentid', 'current_oi_rate',
                    'second_oi_rate', 'third_oi_rate', 'current_oi', 'second_oi', 'third_oi',
                    'current_close', 'second_close', 'third_close', 'current_delist',
                    'second_delist', 'third_delist', 'current_TTM', 'second_TTM', 'third_TTM',
                    'current_OI_speed', 'second_OI_speed', 'third_OI_speed', 'spread_01',
                    'spread_12', 'spread_pct_01', 'spread_pct_12']
    date_list = data['tradingday'].unique()
    date_list = date_list[date_list > '2018-01-01']
    factor = pd.DataFrame()
    for date in date_list:
        result_all = pd.DataFrame()
        for tickers in ConstFutBasic.fut_code_list:
            try:
                result_one_ticker = oi_change_rate_table_one_ticker(tickers, date, data)
                result_all = pd.concat([result_all, result_one_ticker])
            except Exception as e:
                pass
            print(tickers + 'is ok')
        print(date + ' is ok')
        result_all['tradingday'] = date
        factor = pd.concat([factor, result_all])
    factor.to_csv('C:\\Users\\jason.huang\\research\\factor_data\\OI_change\\oi_change_factor.csv')


if __name__ == '__main__':
    tickers_list = ConstFutBasic.fut_code_list

    start_date = '2022-01-01'
    end_date = '2023-02-06'
    file_path = 'C:\\Users\\jason.huang\\research\\scripts_working\\OI_change_rate\\daily\\'

    result1 = oi_change_rate_indicator(tickers_list, start_date, end_date, file_path, daily_index=True)
    oi_change_rate_table(date=end_date)
