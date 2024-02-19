import pandas as pd
import os
import numpy as np

from openpyxl import load_workbook
import matplotlib.pyplot as plt
import datetime
import seaborn as sns

sector_group_name_list = ['cm_cn_group_grains_oilseeds', 'cm_cn_group_livestock', 'cm_cn_group_softs',
                          'cm_cn_group_base_metal', 'cm_cn_group_black', 'cm_cn_group_chemicals', 'cm_cn_group_energy',
                          'cm_cn_group_stock_index', 'cm_cn_group_interest_rate', 'cm_cn_sector_agriculture',
                          'cm_cn_sector_industrials', 'cm_cn_sector_refineries', 'cm_cn_sector_financial', 'cm_cn_all']
fut_code_list = ['A.DCE', 'AL.SHF', 'AP.CZC', 'BU.SHF', 'C.DCE', 'CF.CZC', 'CJ.CZC', 'CS.DCE', 'CU.SHF', 'EB.DCE',
                 'EG.DCE', 'FG.CZC', 'FU.SHF', 'HC.SHF', 'IC.CFE', 'I.DCE', 'IF.CFE', 'IH.CFE', 'J.DCE', 'JD.DCE',
                 'JM.DCE', 'L.DCE', 'LH.DCE', 'MA.CZC', 'M.DCE', 'NI.SHF', 'OI.CZC', 'PB.SHF', 'P.DCE', 'PF.CZC',
                 'PG.DCE', 'PK.CZC', 'PP.DCE', 'RB.SHF', 'RM.CZC', 'RU.SHF', 'SA.CZC', 'SF.CZC', 'SM.CZC', 'SN.SHF',
                 'SP.SHF', 'SR.CZC', 'SS.SHF', 'TA.CZC', 'T.CFE', 'TF.CFE', 'TS.CFE', 'UR.CZC', 'V.DCE', 'Y.DCE',
                 'ZC.CZC', 'ZN.SHF']
mad_param = [0.2, 1, 2, 10]
sign_switch_param = [0, 30, 256, 512, 1024]
factor_list = ['cb_ttm', 'cb_ttm_seasonal', 'cb_diff']

parent_path = 'C:\\Users\\jason.huang\\research\\backtest\\single_factor_test\\cashbasis_front_contract_test\\result_no_signal_delete\\'

excel_save_file_path = 'C:\\Users\\jason.huang\\research\\backtest\\single_factor_test\\cashbasis_front_contract_test\\excel_factor_group\\'

name = 'cm_cn_group_chemicals'
factor = 'cb_ttm'
fee = 0

long_index = False
short_index = False

if fee == 0:
    fee_str = '_no_fee'
else:
    fee_str = ''

# # excel format
# for name in ['cm_cn_group_grains_oilseeds', 'cm_cn_group_softs', 'cm_cn_group_base_metal', 'cm_cn_group_black',
#              'cm_cn_group_chemicals', 'cm_cn_group_livestock', 'cm_cn_group_energy', 'cm_cn_sector_agriculture',
#              'cm_cn_sector_industrials', 'cm_cn_sector_refineries']:
#     if '.' in name:
#         sheet_name = 'cum_pnl' + fee_str
#         sheet_name_indicator = 'Bt_Indicator' + fee_str
#     else:
#         sheet_name = 'sector_cum_pnl' + fee_str
#         sheet_name_indicator = 'sector_Bt_Indicator' + fee_str
#     for factor in factor_list:
#         try:
#             # long short
#             select_name_list_all = []
#             select_name_list_ttm = []
#
#             for file_name in os.listdir(parent_path):
#                 if factor + '_mad' in file_name:
#                     if 'long' not in file_name and 'short' not in file_name:
#                         if 'all' in file_name:
#                             select_name_list_all.append(file_name)
#                         if 'ttm40' in file_name:
#                             select_name_list_ttm.append(file_name)
#             indicator_all = pd.DataFrame()
#             cumpnl_all = pd.DataFrame()
#             for i in range(len(select_name_list_all)):
#                 try:
#                     file_name = select_name_list_all[i]
#                     file_location = parent_path + file_name
#
#                     indicator_all.loc[i, 'mad'] = float(file_name[:-5].split('_')[-3][:-10])
#                     indicator_all.loc[i, 'signswitch'] = float(file_name[:-5].split('_')[-2])
#
#                     # cumpnl
#                     data = pd.read_excel(file_location, sheet_name, index_col=0)
#                     data = data[name]
#                     data = pd.DataFrame(data)
#                     data.columns = [file_name[:-5]]
#
#                     indicator_all.loc[i, 'PNL'] = round(data.iloc[-1, 0])
#
#                     data_indicator = pd.read_excel(file_location, sheet_name_indicator, index_col=0)
#                     data_indicator = data_indicator[data_indicator['code'] == name]
#                     data_indicator = np.array(data_indicator['IR'])[0]
#                     indicator_all.loc[i, 'IR'] = round(data_indicator * 15.8, 2)
#
#                     cumpnl_all = pd.concat([cumpnl_all, data], axis=1)
#                 except Exception as e:
#                     pass
#
#             indicator_ttm = pd.DataFrame()
#             cumpnl_ttm = pd.DataFrame()
#             for i in range(len(select_name_list_ttm)):
#                 try:
#                     file_name = select_name_list_ttm[i]
#                     file_location = parent_path + file_name
#
#                     indicator_ttm.loc[i, 'mad'] = float(file_name[:-5].split('_')[-3][:-10])
#                     indicator_ttm.loc[i, 'signswitch'] = float(file_name[:-5].split('_')[-2])
#
#                     # cumpnl
#                     data = pd.read_excel(file_location, sheet_name, index_col=0)
#                     data = data[name]
#                     data = pd.DataFrame(data)
#                     data.columns = [file_name[:-5]]
#
#                     indicator_ttm.loc[i, 'PNL'] = round(data.iloc[-1, 0])
#
#                     data_indicator = pd.read_excel(file_location, sheet_name_indicator, index_col=0)
#                     data_indicator = data_indicator[data_indicator['code'] == name]
#                     data_indicator = np.array(data_indicator['IR'])[0]
#                     indicator_ttm.loc[i, 'IR'] = round(data_indicator * 15.8, 2)
#
#                     cumpnl_ttm = pd.concat([cumpnl_ttm, data], axis=1)
#                 except Exception as e:
#                     pass
#
#             indicator_all_pnl = indicator_all.pivot_table(index='mad', columns='signswitch', values='PNL')
#             indicator_all_ir = indicator_all.pivot_table(index='mad', columns='signswitch', values='IR')
#             indicator_ttm_pnl = indicator_ttm.pivot_table(index='mad', columns='signswitch', values='PNL')
#             indicator_ttm_ir = indicator_ttm.pivot_table(index='mad', columns='signswitch', values='IR')
#             cumpnl_all = cumpnl_all.T.sort_values(cumpnl_all.T.columns[-1], ascending=False).T
#             cumpnl_ttm = cumpnl_ttm.T.sort_values(cumpnl_ttm.T.columns[-1], ascending=False).T
#
#             excel_save_location = excel_save_file_path + name + '_' + factor + '_longshort.xlsx'
#
#             df_list = [indicator_all_pnl, indicator_all_ir, indicator_ttm_pnl, indicator_ttm_ir, cumpnl_all, cumpnl_ttm]
#             sheet_list = ['indicator_all_pnl', 'indicator_all_ir', 'indicator_ttm_pnl', 'indicator_ttm_ir',
#                           'cumpnl_all',
#                           'cumpnl_ttm']
#             writer = pd.ExcelWriter(excel_save_location, engine='xlsxwriter')
#             for dataframe, sheet in zip(df_list, sheet_list):
#                 dataframe.to_excel(writer, sheet_name=sheet, startrow=0, startcol=0)
#             writer.save()
#
#             format_file_location = r'C:\Users\jason.huang\research\backtest\single_factor_test\cashbasis_front_contract_test\excel_format\format.xlsx'
#             workbook = load_workbook(filename=format_file_location)
#             for f in ['indicator_all_pnl', 'indicator_all_ir', 'indicator_ttm_pnl', 'indicator_ttm_ir', 'cumpnl_all',
#                       'cumpnl_ttm']:
#                 sheet = workbook[f]
#                 df = eval(f)
#                 for j in range(len(df.columns)):
#                     sheet.cell(row=1, column=j + 2).value = df.columns[j]
#                     for i in range(len(df)):
#                         sheet.cell(row=i + 2, column=j + 2).value = df.iloc[i, j]
#                         sheet.cell(row=i + 2, column=1).value = df.index[i]
#
#             workbook.save(
#                 filename="C:\\Users\\jason.huang\\research\\backtest\\single_factor_test\\cashbasis_front_contract_test\\excel_format\\" + name + '_' + factor + "_longshort.xlsx")
#
#             print(name + '   ' + factor + 'long short is ok')
#         except:
#             print(name + '   ' + factor + 'long short is not ok')
#
#         try:
#             # long
#             select_name_list_all = []
#             select_name_list_ttm = []
#
#             for file_name in os.listdir(parent_path):
#                 if factor + '_long_mad' in file_name:
#                     if 'all' in file_name:
#                         select_name_list_all.append(file_name)
#                     if 'ttm40' in file_name:
#                         select_name_list_ttm.append(file_name)
#             indicator_all = pd.DataFrame()
#             cumpnl_all = pd.DataFrame()
#             for i in range(len(select_name_list_all)):
#                 try:
#                     file_name = select_name_list_all[i]
#                     file_location = parent_path + file_name
#
#                     indicator_all.loc[i, 'mad'] = float(file_name[:-5].split('_')[-3][:-10])
#                     indicator_all.loc[i, 'signswitch'] = float(file_name[:-5].split('_')[-2])
#
#                     # cumpnl
#                     data = pd.read_excel(file_location, sheet_name, index_col=0)
#                     data = data[name]
#                     data = pd.DataFrame(data)
#                     data.columns = [file_name[:-5]]
#
#                     indicator_all.loc[i, 'PNL'] = round(data.iloc[-1, 0])
#
#                     data_indicator = pd.read_excel(file_location, sheet_name_indicator, index_col=0)
#                     data_indicator = data_indicator[data_indicator['code'] == name]
#                     data_indicator = np.array(data_indicator['IR'])[0]
#                     indicator_all.loc[i, 'IR'] = round(data_indicator * 15.8, 2)
#
#                     cumpnl_all = pd.concat([cumpnl_all, data], axis=1)
#                 except Exception as e:
#                     pass
#
#             indicator_ttm = pd.DataFrame()
#             cumpnl_ttm = pd.DataFrame()
#             for i in range(len(select_name_list_ttm)):
#                 try:
#                     file_name = select_name_list_ttm[i]
#                     file_location = parent_path + file_name
#
#                     indicator_ttm.loc[i, 'mad'] = float(file_name[:-5].split('_')[-3][:-10])
#                     indicator_ttm.loc[i, 'signswitch'] = float(file_name[:-5].split('_')[-2])
#
#                     # cumpnl
#                     data = pd.read_excel(file_location, sheet_name, index_col=0)
#                     data = data[name]
#                     data = pd.DataFrame(data)
#                     data.columns = [file_name[:-5]]
#
#                     indicator_ttm.loc[i, 'PNL'] = round(data.iloc[-1, 0])
#
#                     data_indicator = pd.read_excel(file_location, sheet_name_indicator, index_col=0)
#                     data_indicator = data_indicator[data_indicator['code'] == name]
#                     data_indicator = np.array(data_indicator['IR'])[0]
#                     indicator_ttm.loc[i, 'IR'] = round(data_indicator * 15.8, 2)
#
#                     cumpnl_ttm = pd.concat([cumpnl_ttm, data], axis=1)
#                 except Exception as e:
#                     pass
#
#             indicator_all_pnl = indicator_all.pivot_table(index='mad', columns='signswitch', values='PNL')
#             indicator_all_ir = indicator_all.pivot_table(index='mad', columns='signswitch', values='IR')
#             indicator_ttm_pnl = indicator_ttm.pivot_table(index='mad', columns='signswitch', values='PNL')
#             indicator_ttm_ir = indicator_ttm.pivot_table(index='mad', columns='signswitch', values='IR')
#             cumpnl_all = cumpnl_all.T.sort_values(cumpnl_all.T.columns[-1], ascending=False).T
#             cumpnl_ttm = cumpnl_ttm.T.sort_values(cumpnl_ttm.T.columns[-1], ascending=False).T
#
#             excel_save_location = excel_save_file_path + name + '_' + factor + '_long.xlsx'
#
#             df_list = [indicator_all_pnl, indicator_all_ir, indicator_ttm_pnl, indicator_ttm_ir, cumpnl_all, cumpnl_ttm]
#             sheet_list = ['indicator_all_pnl', 'indicator_all_ir', 'indicator_ttm_pnl', 'indicator_ttm_ir',
#                           'cumpnl_all',
#                           'cumpnl_ttm']
#             writer = pd.ExcelWriter(excel_save_location, engine='xlsxwriter')
#             for dataframe, sheet in zip(df_list, sheet_list):
#                 dataframe.to_excel(writer, sheet_name=sheet, startrow=0, startcol=0)
#             writer.save()
#
#             format_file_location = r'C:\Users\jason.huang\research\backtest\single_factor_test\cashbasis_front_contract_test\excel_format\format.xlsx'
#             workbook = load_workbook(filename=format_file_location)
#             for f in ['indicator_all_pnl', 'indicator_all_ir', 'indicator_ttm_pnl', 'indicator_ttm_ir', 'cumpnl_all',
#                       'cumpnl_ttm']:
#                 sheet = workbook[f]
#                 df = eval(f)
#                 for j in range(len(df.columns)):
#                     sheet.cell(row=1, column=j + 2).value = df.columns[j]
#                     for i in range(len(df)):
#                         sheet.cell(row=i + 2, column=j + 2).value = df.iloc[i, j]
#                         sheet.cell(row=i + 2, column=1).value = df.index[i]
#
#             workbook.save(
#                 filename="C:\\Users\\jason.huang\\research\\backtest\\single_factor_test\\cashbasis_front_contract_test\\excel_format\\" + name + '_' + factor + "_long.xlsx")
#
#             print(name + '   ' + factor + 'long is ok')
#         except Exception as e:
#             print(name + '   ' + factor + 'long is not ok')
#
#         try:
#             # short
#             select_name_list_all = []
#             select_name_list_ttm = []
#
#             for file_name in os.listdir(parent_path):
#                 if factor + '_short_mad' in file_name:
#                     if 'all' in file_name:
#                         select_name_list_all.append(file_name)
#                     if 'ttm40' in file_name:
#                         select_name_list_ttm.append(file_name)
#             indicator_all = pd.DataFrame()
#             cumpnl_all = pd.DataFrame()
#             for i in range(len(select_name_list_all)):
#                 try:
#                     file_name = select_name_list_all[i]
#                     file_location = parent_path + file_name
#
#                     indicator_all.loc[i, 'mad'] = float(file_name[:-5].split('_')[-3][:-10])
#                     indicator_all.loc[i, 'signswitch'] = float(file_name[:-5].split('_')[-2])
#
#                     # cumpnl
#                     data = pd.read_excel(file_location, sheet_name, index_col=0)
#                     data = data[name]
#                     data = pd.DataFrame(data)
#                     data.columns = [file_name[:-5]]
#
#                     indicator_all.loc[i, 'PNL'] = round(data.iloc[-1, 0])
#
#                     data_indicator = pd.read_excel(file_location, sheet_name_indicator, index_col=0)
#                     data_indicator = data_indicator[data_indicator['code'] == name]
#                     data_indicator = np.array(data_indicator['IR'])[0]
#                     indicator_all.loc[i, 'IR'] = round(data_indicator * 15.8, 2)
#
#                     cumpnl_all = pd.concat([cumpnl_all, data], axis=1)
#                 except Exception as e:
#                     pass
#
#             indicator_ttm = pd.DataFrame()
#             cumpnl_ttm = pd.DataFrame()
#             for i in range(len(select_name_list_ttm)):
#                 try:
#                     file_name = select_name_list_ttm[i]
#                     file_location = parent_path + file_name
#
#                     indicator_ttm.loc[i, 'mad'] = float(file_name[:-5].split('_')[-3][:-10])
#                     indicator_ttm.loc[i, 'signswitch'] = float(file_name[:-5].split('_')[-2])
#
#                     # cumpnl
#                     data = pd.read_excel(file_location, sheet_name, index_col=0)
#                     data = data[name]
#                     data = pd.DataFrame(data)
#                     data.columns = [file_name[:-5]]
#
#                     indicator_ttm.loc[i, 'PNL'] = round(data.iloc[-1, 0])
#
#                     data_indicator = pd.read_excel(file_location, sheet_name_indicator, index_col=0)
#                     data_indicator = data_indicator[data_indicator['code'] == name]
#                     data_indicator = np.array(data_indicator['IR'])[0]
#                     indicator_ttm.loc[i, 'IR'] = round(data_indicator * 15.8, 2)
#
#                     cumpnl_ttm = pd.concat([cumpnl_ttm, data], axis=1)
#                 except Exception as e:
#                     pass
#
#             indicator_all_pnl = indicator_all.pivot_table(index='mad', columns='signswitch', values='PNL')
#             indicator_all_ir = indicator_all.pivot_table(index='mad', columns='signswitch', values='IR')
#             indicator_ttm_pnl = indicator_ttm.pivot_table(index='mad', columns='signswitch', values='PNL')
#             indicator_ttm_ir = indicator_ttm.pivot_table(index='mad', columns='signswitch', values='IR')
#             cumpnl_all = cumpnl_all.T.sort_values(cumpnl_all.T.columns[-1], ascending=False).T
#             cumpnl_ttm = cumpnl_ttm.T.sort_values(cumpnl_ttm.T.columns[-1], ascending=False).T
#
#             excel_save_location = excel_save_file_path + name + '_' + factor + '_short.xlsx'
#
#             df_list = [indicator_all_pnl, indicator_all_ir, indicator_ttm_pnl, indicator_ttm_ir, cumpnl_all, cumpnl_ttm]
#             sheet_list = ['indicator_all_pnl', 'indicator_all_ir', 'indicator_ttm_pnl', 'indicator_ttm_ir',
#                           'cumpnl_all',
#                           'cumpnl_ttm']
#             writer = pd.ExcelWriter(excel_save_location, engine='xlsxwriter')
#             for dataframe, sheet in zip(df_list, sheet_list):
#                 dataframe.to_excel(writer, sheet_name=sheet, startrow=0, startcol=0)
#             writer.save()
#
#             format_file_location = r'C:\Users\jason.huang\research\backtest\single_factor_test\cashbasis_front_contract_test\excel_format\format.xlsx'
#             workbook = load_workbook(filename=format_file_location)
#             for f in ['indicator_all_pnl', 'indicator_all_ir', 'indicator_ttm_pnl', 'indicator_ttm_ir', 'cumpnl_all',
#                       'cumpnl_ttm']:
#                 sheet = workbook[f]
#                 df = eval(f)
#                 for j in range(len(df.columns)):
#                     sheet.cell(row=1, column=j + 2).value = df.columns[j]
#                     for i in range(len(df)):
#                         sheet.cell(row=i + 2, column=j + 2).value = df.iloc[i, j]
#                         sheet.cell(row=i + 2, column=1).value = df.index[i]
#
#             workbook.save(
#                 filename="C:\\Users\\jason.huang\\research\\backtest\\single_factor_test\\cashbasis_front_contract_test\\excel_format\\" + name + '_' + factor + "_short.xlsx")
#
#             print(name + '   ' + factor + 'short is ok')
#         except Exception as e:
#             print(name + '   ' + factor + 'short is not ok')

# excel format version 2

for name in ['cm_cn_group_grains_oilseeds', 'cm_cn_group_softs', 'cm_cn_group_base_metal', 'cm_cn_group_black',
             'cm_cn_group_chemicals', 'cm_cn_group_livestock', 'cm_cn_group_energy', 'cm_cn_sector_agriculture',
             'cm_cn_sector_industrials', 'cm_cn_sector_refineries']:
    if '.' in name:
        sheet_name = 'cum_pnl' + fee_str
        sheet_name_indicator = 'Bt_Indicator' + fee_str
    else:
        sheet_name = 'sector_cum_pnl' + fee_str
        sheet_name_indicator = 'sector_Bt_Indicator' + fee_str
    for factor in factor_list:
        try:
            # long short
            select_name_list_all = []
            select_name_list_ttm = []

            for file_name in os.listdir(parent_path):
                if factor + '_mad' in file_name:
                    if 'long' not in file_name and 'short' not in file_name:
                        if 'all' in file_name:
                            select_name_list_all.append(file_name)
                        if 'ttm40' in file_name:
                            select_name_list_ttm.append(file_name)
            indicator_all = pd.DataFrame()
            cumpnl_all = pd.DataFrame()
            for i in range(len(select_name_list_all)):
                try:
                    file_name = select_name_list_all[i]
                    file_location = parent_path + file_name

                    indicator_all.loc[i, 'mad'] = float(file_name[:-5].split('_')[-3][:-10])
                    indicator_all.loc[i, 'signswitch'] = float(file_name[:-5].split('_')[-2])

                    # cumpnl
                    data = pd.read_excel(file_location, sheet_name, index_col=0)
                    data = data[name]
                    data = pd.DataFrame(data)
                    data.columns = [file_name[:-5]]

                    indicator_all.loc[i, 'PNL'] = round(data.iloc[-1, 0])

                    data_indicator = pd.read_excel(file_location, sheet_name_indicator, index_col=0)
                    data_indicator = data_indicator[data_indicator['code'] == name]
                    data_indicator = np.array(data_indicator['IR'])[0]
                    indicator_all.loc[i, 'IR'] = round(data_indicator * 15.8, 2)

                    cumpnl_all = pd.concat([cumpnl_all, data], axis=1)
                except Exception as e:
                    pass

            indicator_ttm = pd.DataFrame()
            cumpnl_ttm = pd.DataFrame()
            for i in range(len(select_name_list_ttm)):
                try:
                    file_name = select_name_list_ttm[i]
                    file_location = parent_path + file_name

                    indicator_ttm.loc[i, 'mad'] = float(file_name[:-5].split('_')[-3][:-10])
                    indicator_ttm.loc[i, 'signswitch'] = float(file_name[:-5].split('_')[-2])

                    # cumpnl
                    data = pd.read_excel(file_location, sheet_name, index_col=0)
                    data = data[name]
                    data = pd.DataFrame(data)
                    data.columns = [file_name[:-5]]

                    indicator_ttm.loc[i, 'PNL'] = round(data.iloc[-1, 0])

                    data_indicator = pd.read_excel(file_location, sheet_name_indicator, index_col=0)
                    data_indicator = data_indicator[data_indicator['code'] == name]
                    data_indicator = np.array(data_indicator['IR'])[0]
                    indicator_ttm.loc[i, 'IR'] = round(data_indicator * 15.8, 2)

                    cumpnl_ttm = pd.concat([cumpnl_ttm, data], axis=1)
                except Exception as e:
                    pass

            indicator_all_pnl = indicator_all.pivot_table(index='mad', columns='signswitch', values='PNL')
            indicator_all_ir = indicator_all.pivot_table(index='mad', columns='signswitch', values='IR')
            indicator_ttm_pnl = indicator_ttm.pivot_table(index='mad', columns='signswitch', values='PNL')
            indicator_ttm_ir = indicator_ttm.pivot_table(index='mad', columns='signswitch', values='IR')
            cumpnl_all = cumpnl_all.T.sort_values(cumpnl_all.T.columns[-1], ascending=False).T
            cumpnl_ttm = cumpnl_ttm.T.sort_values(cumpnl_ttm.T.columns[-1], ascending=False).T

            format_file_location = r'C:\Users\jason.huang\research\backtest\single_factor_test\cashbasis_front_contract_test\version_2_excel_format\format_indicator.xlsx'
            workbook = load_workbook(filename=format_file_location)
            for f_i in range(len(['indicator_all_pnl', 'indicator_ttm_pnl', 'indicator_all_ir', 'indicator_ttm_ir'])):
                f = ['indicator_all_pnl', 'indicator_ttm_pnl', 'indicator_all_ir', 'indicator_ttm_ir'][f_i]
                sheet = workbook['Indicator']
                df = eval(f)
                for j in range(len(df.columns)):
                    for i in range(len(df)):
                        sheet.cell(row=i + 2 + f_i * 7, column=j + 2).value = df.iloc[i, j]
            workbook.save(
                filename="C:\\Users\\jason.huang\\research\\backtest\\single_factor_test\\cashbasis_front_contract_test\\version_2_excel_format\\" + name + '_' + factor + "_indicator.xlsx")

            format_file_location = r'C:\Users\jason.huang\research\backtest\single_factor_test\cashbasis_front_contract_test\version_2_excel_format\format_pnl_graph.xlsx'
            workbook = load_workbook(filename=format_file_location)
            for f in ['cumpnl_all', 'cumpnl_ttm']:
                sheet = workbook[f]
                df = eval(f)
                for j in range(len(df.columns)):
                    sheet.cell(row=1, column=j + 2).value = df.columns[j]
                    for i in range(len(df)):
                        sheet.cell(row=i + 2, column=j + 2).value = df.iloc[i, j]
                        sheet.cell(row=i + 2, column=1).value = df.index[i]
            workbook.save(
                filename="C:\\Users\\jason.huang\\research\\backtest\\single_factor_test\\cashbasis_front_contract_test\\version_2_excel_format\\" + name + '_' + factor + "_longshort.xlsx")

            print(name + '   ' + factor + 'long short is ok')
        except:
            print(name + '   ' + factor + 'long short is not ok')

        try:
            # long
            select_name_list_all = []
            select_name_list_ttm = []

            for file_name in os.listdir(parent_path):
                if factor + '_long_mad' in file_name:
                    if 'all' in file_name:
                        select_name_list_all.append(file_name)
                    if 'ttm40' in file_name:
                        select_name_list_ttm.append(file_name)
            indicator_all = pd.DataFrame()
            cumpnl_all = pd.DataFrame()
            for i in range(len(select_name_list_all)):
                try:
                    file_name = select_name_list_all[i]
                    file_location = parent_path + file_name

                    indicator_all.loc[i, 'mad'] = float(file_name[:-5].split('_')[-3][:-10])
                    indicator_all.loc[i, 'signswitch'] = float(file_name[:-5].split('_')[-2])

                    # cumpnl
                    data = pd.read_excel(file_location, sheet_name, index_col=0)
                    data = data[name]
                    data = pd.DataFrame(data)
                    data.columns = [file_name[:-5]]

                    indicator_all.loc[i, 'PNL'] = round(data.iloc[-1, 0])

                    data_indicator = pd.read_excel(file_location, sheet_name_indicator, index_col=0)
                    data_indicator = data_indicator[data_indicator['code'] == name]
                    data_indicator = np.array(data_indicator['IR'])[0]
                    indicator_all.loc[i, 'IR'] = round(data_indicator * 15.8, 2)

                    cumpnl_all = pd.concat([cumpnl_all, data], axis=1)
                except Exception as e:
                    pass

            indicator_ttm = pd.DataFrame()
            cumpnl_ttm = pd.DataFrame()
            for i in range(len(select_name_list_ttm)):
                try:
                    file_name = select_name_list_ttm[i]
                    file_location = parent_path + file_name

                    indicator_ttm.loc[i, 'mad'] = float(file_name[:-5].split('_')[-3][:-10])
                    indicator_ttm.loc[i, 'signswitch'] = float(file_name[:-5].split('_')[-2])

                    # cumpnl
                    data = pd.read_excel(file_location, sheet_name, index_col=0)
                    data = data[name]
                    data = pd.DataFrame(data)
                    data.columns = [file_name[:-5]]

                    indicator_ttm.loc[i, 'PNL'] = round(data.iloc[-1, 0])

                    data_indicator = pd.read_excel(file_location, sheet_name_indicator, index_col=0)
                    data_indicator = data_indicator[data_indicator['code'] == name]
                    data_indicator = np.array(data_indicator['IR'])[0]
                    indicator_ttm.loc[i, 'IR'] = round(data_indicator * 15.8, 2)

                    cumpnl_ttm = pd.concat([cumpnl_ttm, data], axis=1)
                except Exception as e:
                    pass

            indicator_all_pnl = indicator_all.pivot_table(index='mad', columns='signswitch', values='PNL')
            indicator_all_ir = indicator_all.pivot_table(index='mad', columns='signswitch', values='IR')
            indicator_ttm_pnl = indicator_ttm.pivot_table(index='mad', columns='signswitch', values='PNL')
            indicator_ttm_ir = indicator_ttm.pivot_table(index='mad', columns='signswitch', values='IR')
            cumpnl_all = cumpnl_all.T.sort_values(cumpnl_all.T.columns[-1], ascending=False).T
            cumpnl_ttm = cumpnl_ttm.T.sort_values(cumpnl_ttm.T.columns[-1], ascending=False).T

            format_file_location = "C:\\Users\\jason.huang\\research\\backtest\\single_factor_test\\cashbasis_front_contract_test\\version_2_excel_format\\" + name + '_' + factor + "_indicator.xlsx"
            workbook = load_workbook(filename=format_file_location)
            for f_i in range(len(['indicator_all_pnl', 'indicator_ttm_pnl', 'indicator_all_ir', 'indicator_ttm_ir'])):
                f = ['indicator_all_pnl', 'indicator_ttm_pnl', 'indicator_all_ir', 'indicator_ttm_ir'][f_i]
                sheet = workbook['Indicator']
                df = eval(f)
                for j in range(len(df.columns)):
                    for i in range(len(df)):
                        sheet.cell(row=i + 2 + f_i * 7, column=j + 2 + 7).value = df.iloc[i, j]

            workbook.save(
                filename="C:\\Users\\jason.huang\\research\\backtest\\single_factor_test\\cashbasis_front_contract_test\\version_2_excel_format\\" + name + '_' + factor + "_indicator.xlsx")

            format_file_location = r'C:\Users\jason.huang\research\backtest\single_factor_test\cashbasis_front_contract_test\version_2_excel_format\format_pnl_graph.xlsx'
            workbook = load_workbook(filename=format_file_location)
            for f in ['cumpnl_all', 'cumpnl_ttm']:
                sheet = workbook[f]
                df = eval(f)
                for j in range(len(df.columns)):
                    sheet.cell(row=1, column=j + 2).value = df.columns[j]
                    for i in range(len(df)):
                        sheet.cell(row=i + 2, column=j + 2).value = df.iloc[i, j]
                        sheet.cell(row=i + 2, column=1).value = df.index[i]
            workbook.save(
                filename="C:\\Users\\jason.huang\\research\\backtest\\single_factor_test\\cashbasis_front_contract_test\\version_2_excel_format\\" + name + '_' + factor + "_long.xlsx")

            print(name + '   ' + factor + 'long is ok')
        except Exception as e:
            print(name + '   ' + factor + 'long is not ok')

        try:
            # short
            select_name_list_all = []
            select_name_list_ttm = []

            for file_name in os.listdir(parent_path):
                if factor + '_short_mad' in file_name:
                    if 'all' in file_name:
                        select_name_list_all.append(file_name)
                    if 'ttm40' in file_name:
                        select_name_list_ttm.append(file_name)
            indicator_all = pd.DataFrame()
            cumpnl_all = pd.DataFrame()
            for i in range(len(select_name_list_all)):
                try:
                    file_name = select_name_list_all[i]
                    file_location = parent_path + file_name

                    indicator_all.loc[i, 'mad'] = float(file_name[:-5].split('_')[-3][:-10])
                    indicator_all.loc[i, 'signswitch'] = float(file_name[:-5].split('_')[-2])

                    # cumpnl
                    data = pd.read_excel(file_location, sheet_name, index_col=0)
                    data = data[name]
                    data = pd.DataFrame(data)
                    data.columns = [file_name[:-5]]

                    indicator_all.loc[i, 'PNL'] = round(data.iloc[-1, 0])

                    data_indicator = pd.read_excel(file_location, sheet_name_indicator, index_col=0)
                    data_indicator = data_indicator[data_indicator['code'] == name]
                    data_indicator = np.array(data_indicator['IR'])[0]
                    indicator_all.loc[i, 'IR'] = round(data_indicator * 15.8, 2)

                    cumpnl_all = pd.concat([cumpnl_all, data], axis=1)
                except Exception as e:
                    pass

            indicator_ttm = pd.DataFrame()
            cumpnl_ttm = pd.DataFrame()
            for i in range(len(select_name_list_ttm)):
                try:
                    file_name = select_name_list_ttm[i]
                    file_location = parent_path + file_name

                    indicator_ttm.loc[i, 'mad'] = float(file_name[:-5].split('_')[-3][:-10])
                    indicator_ttm.loc[i, 'signswitch'] = float(file_name[:-5].split('_')[-2])

                    # cumpnl
                    data = pd.read_excel(file_location, sheet_name, index_col=0)
                    data = data[name]
                    data = pd.DataFrame(data)
                    data.columns = [file_name[:-5]]

                    indicator_ttm.loc[i, 'PNL'] = round(data.iloc[-1, 0])

                    data_indicator = pd.read_excel(file_location, sheet_name_indicator, index_col=0)
                    data_indicator = data_indicator[data_indicator['code'] == name]
                    data_indicator = np.array(data_indicator['IR'])[0]
                    indicator_ttm.loc[i, 'IR'] = round(data_indicator * 15.8, 2)

                    cumpnl_ttm = pd.concat([cumpnl_ttm, data], axis=1)
                except Exception as e:
                    pass

            indicator_all_pnl = indicator_all.pivot_table(index='mad', columns='signswitch', values='PNL')
            indicator_all_ir = indicator_all.pivot_table(index='mad', columns='signswitch', values='IR')
            indicator_ttm_pnl = indicator_ttm.pivot_table(index='mad', columns='signswitch', values='PNL')
            indicator_ttm_ir = indicator_ttm.pivot_table(index='mad', columns='signswitch', values='IR')
            cumpnl_all = cumpnl_all.T.sort_values(cumpnl_all.T.columns[-1], ascending=False).T
            cumpnl_ttm = cumpnl_ttm.T.sort_values(cumpnl_ttm.T.columns[-1], ascending=False).T

            format_file_location = "C:\\Users\\jason.huang\\research\\backtest\\single_factor_test\\cashbasis_front_contract_test\\version_2_excel_format\\" + name + '_' + factor + "_indicator.xlsx"
            workbook = load_workbook(filename=format_file_location)
            for f_i in range(len(['indicator_all_pnl', 'indicator_ttm_pnl', 'indicator_all_ir', 'indicator_ttm_ir'])):
                f = ['indicator_all_pnl', 'indicator_ttm_pnl', 'indicator_all_ir', 'indicator_ttm_ir'][f_i]
                sheet = workbook['Indicator']
                df = eval(f)
                for j in range(len(df.columns)):
                    for i in range(len(df)):
                        sheet.cell(row=i + 2 + f_i * 7, column=j + 2 + 7 + 7).value = df.iloc[i, j]

            workbook.save(
                filename="C:\\Users\\jason.huang\\research\\backtest\\single_factor_test\\cashbasis_front_contract_test\\version_2_excel_format\\" + name + '_' + factor + "_indicator.xlsx")

            format_file_location = r'C:\Users\jason.huang\research\backtest\single_factor_test\cashbasis_front_contract_test\version_2_excel_format\format_pnl_graph.xlsx'
            workbook = load_workbook(filename=format_file_location)
            for f in ['cumpnl_all', 'cumpnl_ttm']:
                sheet = workbook[f]
                df = eval(f)
                for j in range(len(df.columns)):
                    sheet.cell(row=1, column=j + 2).value = df.columns[j]
                    for i in range(len(df)):
                        sheet.cell(row=i + 2, column=j + 2).value = df.iloc[i, j]
                        sheet.cell(row=i + 2, column=1).value = df.index[i]
            workbook.save(
                filename="C:\\Users\\jason.huang\\research\\backtest\\single_factor_test\\cashbasis_front_contract_test\\version_2_excel_format\\" + name + '_' + factor + "_short.xlsx")

            print(name + '   ' + factor + 'short is ok')
        except Exception as e:
            print(name + '   ' + factor + 'short is not ok')
